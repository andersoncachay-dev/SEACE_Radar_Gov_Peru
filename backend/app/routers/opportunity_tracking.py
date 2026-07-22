from __future__ import annotations

from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from ..database import get_db
from ..dependencies import get_current_user, require_source_access, source_access_condition
from ..models import Opportunity, OpportunityTracking, OpportunityTrackingStage, User
from ..schemas import (
    CoResponsibleUpdateIn,
    OpportunityTrackingOut,
    OpportunityTrackingSummaryOut,
    QuotationOutcomeIn,
    StageSupportRequestIn,
    StageSupportRequestOut,
    TrackingDateRefreshStatusOut,
    TrackingExcelExportIn,
    TrackingStageAreasUpdate,
    TrackingStageAssigneesUpdate,
    TrackingStageOut,
    TrackingStageUpdate,
)
from ..services import tracking_service
from ..services.scheduler_service import tracking_date_refresh_status
from ..services.tracking_date_refresh_service import get_last_date_refresh_result
from ..services.tracking_notification_service import send_stage_support_request

router = APIRouter(prefix="/opportunity-tracking", tags=["tracking"])


@router.get("/date-refresh/status", response_model=TrackingDateRefreshStatusOut)
def date_refresh_status(
    country: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if country not in {"peru", "chile"}:
        raise HTTPException(status_code=404, detail="País no soportado")
    status_payload = tracking_date_refresh_status(country)
    last_run = get_last_date_refresh_result(db, country)
    return TrackingDateRefreshStatusOut(**status_payload, last_run=last_run)


@router.post("/export/xlsx")
def export_tracking_xlsx(
    payload: TrackingExcelExportIn,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Consolidado"
    sheet.freeze_panes = "A2"
    sheet.append(payload.headers)
    for row in payload.rows:
        sheet.append(row)

    header_fill = PatternFill("solid", fgColor="0D4F9D")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.auto_filter.ref = sheet.dimensions
    for index in range(1, len(payload.headers) + 1):
        sheet.column_dimensions[get_column_letter(index)].width = 22
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"seguimiento-{datetime.now().date().isoformat()}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _parse_ids(raw: str | None) -> list[int] | None:
    if not raw:
        return None
    return [int(x) for x in raw.split(",") if x.strip().isdigit()]


def _load_opportunity(db: Session, opportunity_id: int, current_user: User) -> Opportunity:
    opportunity = db.get(Opportunity, opportunity_id)
    if not opportunity:
        raise HTTPException(status_code=404, detail="Proceso no encontrado")
    require_source_access(current_user, opportunity.source)
    return opportunity


def _load_tracking(db: Session, opportunity: Opportunity) -> OpportunityTracking:
    tracking = tracking_service.get_tracking_for_opportunity(db, opportunity.id)
    if not tracking:
        raise HTTPException(status_code=404, detail="La oportunidad no esta en seguimiento")
    return tracking


def _tracking_out(db: Session, tracking: OpportunityTracking) -> OpportunityTrackingOut:
    payload = OpportunityTrackingOut.model_validate(tracking).model_dump()
    if tracking.started_by_id:
        started_by = db.get(User, tracking.started_by_id)
        payload["started_by_name"] = started_by.full_name if started_by else ""
    if tracking.co_responsible_id:
        co_responsible = db.get(User, tracking.co_responsible_id)
        payload["co_responsible_name"] = co_responsible.full_name if co_responsible else ""
    return OpportunityTrackingOut(**payload)


def _stage_context(db: Session, stage_id: int, current_user: User) -> tuple[OpportunityTrackingStage, OpportunityTracking, Opportunity]:
    stage = db.get(OpportunityTrackingStage, stage_id)
    if not stage:
        raise HTTPException(status_code=404, detail="Etapa no encontrada")
    tracking = db.get(OpportunityTracking, stage.tracking_id)
    opportunity = db.get(Opportunity, tracking.opportunity_id)
    require_source_access(current_user, opportunity.source)
    return stage, tracking, opportunity


@router.get("", response_model=list[OpportunityTrackingSummaryOut])
def list_trackings(
    opportunity_ids: str | None = None,
    country: str | None = None,
    mine_only: bool = False,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    query = select(OpportunityTracking, Opportunity).join(Opportunity, Opportunity.id == OpportunityTracking.opportunity_id)
    condition = source_access_condition(Opportunity.source, current_user)
    if condition is not None:
        query = query.where(condition)
    if country == "chile":
        query = query.where(Opportunity.source.ilike("mercado_publico%"))
    elif country == "peru":
        query = query.where(~Opportunity.source.ilike("mercado_publico%"))
    if mine_only:
        query = query.where(OpportunityTracking.started_by_id == current_user.id)
    query = query.where(OpportunityTracking.status != "retirado")
    ids = _parse_ids(opportunity_ids)
    if ids:
        query = query.where(OpportunityTracking.opportunity_id.in_(ids))
    rows = db.execute(query.order_by(OpportunityTracking.started_at.desc())).all()

    user_ids = {tracking.started_by_id for tracking, _ in rows if tracking.started_by_id}
    user_ids |= {tracking.co_responsible_id for tracking, _ in rows if tracking.co_responsible_id}
    user_names = {u.id: u.full_name for u in db.scalars(select(User).where(User.id.in_(user_ids)))} if user_ids else {}

    return [
        OpportunityTrackingSummaryOut(
            opportunity_id=opportunity.id,
            entity=opportunity.entity,
            nomenclature=opportunity.nomenclature,
            description=opportunity.description,
            source=opportunity.source,
            status=tracking.status,
            current_phase_id=tracking.current_phase_id,
            quotation_outcome=tracking.quotation_outcome,
            publication_date=opportunity.publication_date,
            proposal_deadline=opportunity.proposal_deadline,
            quote_deadline=opportunity.quote_deadline,
            documents_count=opportunity.documents_count,
            requirement_pdf_url=opportunity.requirement_pdf_url,
            started_by_id=tracking.started_by_id,
            started_by_name=user_names.get(tracking.started_by_id, ""),
            co_responsible_id=tracking.co_responsible_id,
            co_responsible_name=user_names.get(tracking.co_responsible_id, ""),
        )
        for tracking, opportunity in rows
    ]


@router.get("/{opportunity_id}", response_model=OpportunityTrackingOut)
def get_tracking(
    opportunity_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    opportunity = _load_opportunity(db, opportunity_id, current_user)
    tracking = _load_tracking(db, opportunity)
    return _tracking_out(db, tracking)


@router.put("/{opportunity_id}/co-responsible", response_model=OpportunityTrackingOut)
def set_co_responsible(
    opportunity_id: int,
    payload: CoResponsibleUpdateIn,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    opportunity = _load_opportunity(db, opportunity_id, current_user)
    tracking = _load_tracking(db, opportunity)
    if current_user.id != tracking.started_by_id:
        raise HTTPException(
            status_code=403,
            detail="Solo el responsable de la oportunidad puede asignar o cambiar al corresponsable",
        )
    if payload.user_id is not None:
        co_responsible = db.get(User, payload.user_id)
        if not co_responsible or not co_responsible.is_active:
            raise HTTPException(status_code=404, detail="Usuario no encontrado")
        require_source_access(co_responsible, opportunity.source)
    tracking.co_responsible_id = payload.user_id
    db.commit()
    db.refresh(tracking)
    return _tracking_out(db, tracking)


@router.post("/{opportunity_id}/withdraw", response_model=OpportunityTrackingOut)
def withdraw_tracking(
    opportunity_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    opportunity = _load_opportunity(db, opportunity_id, current_user)
    tracking = _load_tracking(db, opportunity)
    if current_user.id != tracking.started_by_id:
        raise HTTPException(
            status_code=403,
            detail="Solo el responsable de la oportunidad puede retirarla de seguimiento",
        )
    tracking.status = "retirado"
    db.commit()
    db.refresh(tracking)
    return _tracking_out(db, tracking)


@router.post("/{opportunity_id}/start", response_model=OpportunityTrackingOut, status_code=status.HTTP_201_CREATED)
def start_tracking(
    opportunity_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    opportunity = _load_opportunity(db, opportunity_id, current_user)
    tracking = tracking_service.start_tracking(db, opportunity, current_user)
    return _tracking_out(db, tracking)


@router.patch("/stages/{stage_id}", response_model=TrackingStageOut)
def update_stage(
    stage_id: int,
    payload: TrackingStageUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    stage, _tracking, _opportunity = _stage_context(db, stage_id, current_user)
    updates = payload.model_dump(exclude_unset=True)
    return tracking_service.toggle_stage(db, stage, updates, current_user)


@router.put("/stages/{stage_id}/areas", response_model=TrackingStageOut)
def update_stage_areas(
    stage_id: int,
    payload: TrackingStageAreasUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    stage, _tracking, _opportunity = _stage_context(db, stage_id, current_user)
    return tracking_service.update_stage_areas(db, stage, payload.area_ids)


@router.put("/stages/{stage_id}/assignees", response_model=TrackingStageOut)
def update_stage_assignees(
    stage_id: int,
    payload: TrackingStageAssigneesUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    stage, _tracking, opportunity = _stage_context(db, stage_id, current_user)
    return tracking_service.update_stage_assignees(db, stage, payload.responsible_ids, opportunity, current_user)


@router.post("/stages/{stage_id}/notify", response_model=StageSupportRequestOut)
def notify_stage_support(
    stage_id: int,
    payload: StageSupportRequestIn,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    stage, tracking, opportunity = _stage_context(db, stage_id, current_user)
    owner = db.get(User, tracking.started_by_id) if tracking.started_by_id else None
    owner_name = owner.full_name if owner else current_user.full_name
    sent, failed = send_stage_support_request(db, stage, opportunity, payload.responsible_ids, payload.message, owner_name)
    return StageSupportRequestOut(sent=sent, failed=failed)


@router.post("/{opportunity_id}/quotation-outcome", response_model=OpportunityTrackingOut)
def set_quotation_outcome(
    opportunity_id: int,
    payload: QuotationOutcomeIn,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    opportunity = _load_opportunity(db, opportunity_id, current_user)
    tracking = _load_tracking(db, opportunity)
    tracking = tracking_service.set_quotation_outcome(db, tracking, opportunity, payload.outcome, current_user)
    return _tracking_out(db, tracking)


@router.post("/{opportunity_id}/advance-phase", response_model=OpportunityTrackingOut)
def advance_phase(
    opportunity_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    opportunity = _load_opportunity(db, opportunity_id, current_user)
    tracking = _load_tracking(db, opportunity)
    tracking = tracking_service.advance_phase(db, tracking, opportunity, current_user)
    return _tracking_out(db, tracking)
