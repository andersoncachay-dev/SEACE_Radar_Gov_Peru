from __future__ import annotations

import pandas as pd
import tempfile
from io import BytesIO
from datetime import datetime
from pathlib import Path

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from ..database import get_db
from ..dependencies import get_current_user, require_source_access, source_access_condition
from ..config import settings
from ..models import Opportunity, OpportunitySnapshot, ScrapeRun, User
from ..schemas import OpportunityExcelExportIn, OpportunityImportIn, OpportunityImportResult, OpportunityKeywordArchiveIn, OpportunityKeywordArchiveOut, OpportunityOut, OpportunitySnapshotOut
from src.keyword_matching import contains_complete_phrase
from ..services.ingestion_service import upsert_opportunities
from ..services.notification_service import evaluate_alerts, evaluate_new_opportunity_alerts, send_pending_alerts
from ..services.seace_excel_service import read_seace_export
from ..services.entity_catalog_service import find_entity

router = APIRouter(prefix="/opportunities", tags=["opportunities"])


@router.post("/export/xlsx")
def export_opportunities_xlsx(
    payload: OpportunityExcelExportIn,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    from ..services.scoring_config_service import get_scoring_config
    scoring_config = get_scoring_config(db, payload.country)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Oportunidades"
    sheet.freeze_panes = "A2"
    sheet.append(payload.headers)
    for row in payload.rows:
        sheet.append(row)

    header_fill = PatternFill("solid", fgColor="0D4F9D")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.auto_filter.ref = sheet.dimensions
    widths = [12, 34, 38, 30, 70, 24, 24, 16, 24, 16, 18]
    for index, width in enumerate(widths[: len(payload.headers)], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    definitions = workbook.create_sheet("Definición prioridades")
    definitions.sheet_view.showGridLines = False
    definitions.merge_cells("A1:E1")
    definitions["A1"] = "Definición de prioridades y cálculo del score"
    definitions["A1"].fill = header_fill
    definitions["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    definitions["A1"].alignment = Alignment(horizontal="center", vertical="center")
    definitions.row_dimensions[1].height = 26

    definitions.append([])
    definitions.append(["Prioridad", "Rango de score", "Nivel", "Interpretación"])
    a_min = scoring_config["priority_a_min"]
    b_min = scoring_config["priority_b_min"]
    priority_rows = [
        ["A", f"{a_min} a 100", "Alta", "Oportunidad de atención prioritaria; revisar inmediatamente."],
        ["B", f"{b_min} a {a_min - 1}", "Media", "Oportunidad relevante; revisar y calificar comercialmente."],
        ["C", f"0 a {b_min - 1}", "Baja", "Oportunidad de menor afinidad o urgencia; mantener en seguimiento."],
        ["Objetivo", "100 puntos", "Máximo positivo", "La configuración solo puede guardarse cuando el máximo score positivo alcanzable suma 100."],
    ]
    for row in priority_rows:
        definitions.append(row)

    definitions.append([])
    definitions.append(["Factor evaluado", "Valor considerado", "Puntos", "Condición", "Motivo registrado"])
    conditions = {
        "keyword": ("La descripción, objeto, nomenclatura o área usuaria contiene una keyword core de conectividad/satelital.", "Keyword conectividad/satelital"),
        "target_entity": ("La entidad coincide con el catálogo objetivo o contiene PROVÍAS/MTC.", "Entidad objetivo"),
        "priority_region": ("La región está incluida en el catálogo de regiones priorizadas.", "Región priorizada"),
        "attractive_amount": (f"El monto del proceso es igual o superior a {scoring_config['attractive_amount_min']:,}.", "Monto atractivo"),
        "quick_purchase": ("El origen corresponde a una compra menor a 8 UIT.", "Compra rápida menor a 8 UIT"),
        "queries_and_proposal": ("El proceso está vigente para consultas y propuesta o consulta y cotización.", "Vigente para consultas/cotización"),
        "proposal_only": ("El proceso está vigente únicamente para propuesta o cotización.", "Vigente solo para propuesta/cotización"),
        "evaluation": ("El proceso se encuentra en evaluación.", "En evaluación"),
        "closed": ("El estado comercial indica que el proceso está cerrado.", "Proceso cerrado"),
        "enterprise": ("La descripción contiene requisitos enterprise configurados.", "Requisitos enterprise"),
    }
    score_rows = []
    for key, factor in scoring_config["factors"].items():
        if key in conditions:
            condition, reason = conditions[key]
        else:
            field_labels = {"description": "descripción u objeto", "entity": "entidad", "region": "región", "amount": "monto", "origin": "origen", "status": "estado comercial"}
            field_label = field_labels.get(factor.get("field"), "descripción")
            condition = f"El campo {field_label} coincide con el valor configurado."
            reason = factor["label"]
        points = f"{factor['points']:+d}" if factor["enabled"] else "No aplica"
        considered_value = factor["value"] if factor["enabled"] else "No aplica"
        score_rows.append([factor["label"], considered_value, points, condition if factor["enabled"] else f"No aplica para {payload.country.title()}.", reason if factor["enabled"] else "Desactivado"])
    for row in score_rows:
        definitions.append(row)

    for header_row in (3, 9):
        for cell in definitions[header_row]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    definitions.auto_filter.ref = f"A9:E{definitions.max_row}"
    definitions.freeze_panes = "A4"
    for column, width in {"A": 30, "B": 55, "C": 16, "D": 72, "E": 38}.items():
        definitions.column_dimensions[column].width = width
    for row in definitions.iter_rows(min_row=4):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    note_row = definitions.max_row + 2
    definitions.cell(note_row, 1, "Nota: el score final se limita a un máximo de 100 puntos. Las condiciones aplicables se acumulan.")
    definitions.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=5)
    definitions.cell(note_row, 1).font = Font(italic=True, color="4B5F7A")

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"oportunidades-{datetime.now().date().isoformat()}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _country_for_source(source: str) -> str:
    return "chile" if str(source or "").lower().startswith("mercado_publico") else "peru"


def _archive_key(item: Opportunity) -> str:
    return str(item.nomenclature or item.external_id or "").strip().casefold()


def _enriched_entity_fields(item: Opportunity) -> tuple[str, str]:
    region = str(item.region or "").strip()
    buyer_ruc = str(item.buyer_ruc or "").strip()
    if region and buyer_ruc:
        return region, buyer_ruc
    catalog = find_entity(str(item.entity or ""))
    if not catalog:
        return region, buyer_ruc
    return region or catalog.get("region", ""), buyer_ruc or catalog.get("ruc", "")


def _opportunity_out(item: Opportunity) -> dict:
    payload = OpportunityOut.model_validate(item).model_dump()
    region, buyer_ruc = _enriched_entity_fields(item)
    payload["region"] = region
    payload["buyer_ruc"] = buyer_ruc
    return payload


@router.get("/stats")
def opportunity_stats(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    query = select(Opportunity).where(Opportunity.is_archived.is_(False))
    access_condition = source_access_condition(Opportunity.source, current_user)
    if access_condition is not None:
        query = query.where(access_condition)
    opportunities = list(db.scalars(query).all())
    by_source: dict[str, int] = {}
    by_priority: dict[str, int] = {"A": 0, "B": 0, "C": 0}
    by_region: dict[str, int] = {}
    total_amount = 0.0
    vigentes = 0
    cerrados = 0
    with_ruc = 0
    with_region = 0
    ocds_total = 0
    documents_known = 0
    for item in opportunities:
        by_source[item.source] = by_source.get(item.source, 0) + 1
        by_priority[item.priority] = by_priority.get(item.priority, 0) + 1
        region, buyer_ruc = _enriched_entity_fields(item)
        if region:
            by_region[region] = by_region.get(region, 0) + 1
            with_region += 1
        if buyer_ruc:
            with_ruc += 1
        if str(item.source or "").lower().startswith("oece_ocds"):
            ocds_total += 1
        if int(item.documents_count or 0) > 0 or str(item.requirement_pdf_url or "").strip():
            documents_known += 1
        total_amount += float(item.amount or 0)
        status = str(item.status or "").lower()
        if "vigente" in status:
            vigentes += 1
        if "cerrado" in status:
            cerrados += 1
    return {
        "total": len(opportunities),
        "by_source": by_source,
        "by_priority": by_priority,
        "by_region": by_region,
        "vigentes": vigentes,
        "cerrados": cerrados,
        "total_amount": total_amount,
        "with_ruc": with_ruc,
        "with_region": with_region,
        "ocds_total": ocds_total,
        "documents_known": documents_known,
    }


@router.get("", response_model=list[OpportunityOut])
def list_opportunities(
    source: str | None = None,
    priority: str | None = None,
    status: str | None = None,
    run_id: int | None = None,
    run_ids: str | None = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    query = select(Opportunity).where(Opportunity.is_archived.is_(False))
    access_condition = source_access_condition(Opportunity.source, current_user)
    if access_condition is not None:
        query = query.where(access_condition)
    run_filter_ids: list[int] = []
    if run_id is not None:
        run_filter_ids.append(run_id)
    if run_ids:
        for raw_id in run_ids.split(","):
            try:
                run_filter_ids.append(int(raw_id.strip()))
            except ValueError:
                continue
    if run_filter_ids:
        query = (
            query.join(OpportunitySnapshot, OpportunitySnapshot.opportunity_id == Opportunity.id)
            .where(OpportunitySnapshot.run_id.in_(sorted(set(run_filter_ids))))
            .distinct()
        )
    if source:
        query = query.where(Opportunity.source == source)
    if priority:
        query = query.where(Opportunity.priority == priority)
    if status:
        query = query.where(Opportunity.status == status)
    # Inicio Peru y Chile construyen sus indicadores, mapa y cobertura sobre esta
    # colección. Un límite global dejaba fuera fuentes menos recientes cuando
    # otra fuente (por ejemplo OCDS) ocupaba primero todo el cupo.
    query = query.order_by(Opportunity.updated_at.desc())
    return [_opportunity_out(item) for item in db.scalars(query).all()]


@router.get("/archived", response_model=list[OpportunityOut])
def list_archived_opportunities(
    country: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    normalized_country = country.strip().lower()
    if normalized_country not in {"peru", "chile"}:
        raise HTTPException(status_code=422, detail="El país debe ser Peru o Chile")
    require_source_access(current_user, "mercado_publico_lmp_gc" if normalized_country == "chile" else "oece_ocds")
    query = (
        select(Opportunity)
        .where(
            Opportunity.is_archived.is_(True),
            Opportunity.archive_country == normalized_country,
        )
        .order_by(Opportunity.archived_at.desc(), Opportunity.updated_at.desc())
    )
    return [_opportunity_out(item) for item in db.scalars(query).all()]


@router.post("/archive-by-keyword", response_model=OpportunityKeywordArchiveOut)
def archive_opportunities_by_keyword(
    payload: OpportunityKeywordArchiveIn,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    country = payload.country.strip().lower()
    keyword = payload.keyword.strip()
    remaining_keywords = [item.strip() for item in payload.remaining_keywords if item.strip()]
    if country not in {"peru", "chile"}:
        raise HTTPException(status_code=422, detail="El país debe ser Perú o Chile")
    reference_source = "mercado_publico_lmp_gc" if country == "chile" else "oece_ocds"
    require_source_access(current_user, reference_source)
    candidates = db.scalars(select(Opportunity).where(Opportunity.is_archived.is_(False))).all()
    archived_ids: list[int] = []
    archived_at = datetime.utcnow()
    for opportunity in candidates:
        if _country_for_source(opportunity.source) != country:
            continue
        searchable_text = " ".join((
            str(opportunity.nomenclature or ""),
            str(opportunity.description or ""),
            str(opportunity.entity or ""),
            str(opportunity.object_type or ""),
        ))
        if not contains_complete_phrase(searchable_text, keyword):
            continue
        if any(contains_complete_phrase(searchable_text, other_keyword) for other_keyword in remaining_keywords):
            continue
        key = _archive_key(opportunity)
        if not key:
            continue
        opportunity.is_archived = True
        opportunity.archived_at = archived_at
        opportunity.archived_by_id = current_user.id
        opportunity.archive_country = country
        opportunity.archive_key = key
        archived_ids.append(opportunity.id)
    db.commit()
    return {"archived": len(archived_ids), "opportunity_ids": archived_ids}


@router.post("/{opportunity_id}/archive", response_model=OpportunityOut)
def archive_opportunity(
    opportunity_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    opportunity = db.get(Opportunity, opportunity_id)
    if not opportunity:
        raise HTTPException(status_code=404, detail="Proceso no encontrado")
    require_source_access(current_user, opportunity.source)
    key = _archive_key(opportunity)
    if not key:
        raise HTTPException(status_code=422, detail="El proceso no tiene una nomenclatura válida para archivarlo")
    opportunity.is_archived = True
    opportunity.archived_at = datetime.utcnow()
    opportunity.archived_by_id = current_user.id
    opportunity.archive_country = _country_for_source(opportunity.source)
    opportunity.archive_key = key
    db.commit()
    db.refresh(opportunity)
    return _opportunity_out(opportunity)


@router.post("/{opportunity_id}/restore", response_model=OpportunityOut)
def restore_opportunity(
    opportunity_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    opportunity = db.get(Opportunity, opportunity_id)
    if not opportunity:
        raise HTTPException(status_code=404, detail="Proceso no encontrado")
    require_source_access(current_user, opportunity.source)
    opportunity.is_archived = False
    opportunity.archived_at = None
    opportunity.archived_by_id = None
    opportunity.archive_country = ""
    opportunity.archive_key = ""
    db.commit()
    db.refresh(opportunity)
    return _opportunity_out(opportunity)


@router.post("/import", response_model=OpportunityImportResult)
def import_opportunities(
    payload: OpportunityImportIn,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    require_source_access(current_user, payload.source)
    df = pd.DataFrame(payload.rows or [])
    count = upsert_opportunities(db, df, payload.source)
    return {"imported": count}


@router.post("/import-seace-excel", response_model=OpportunityImportResult)
async def import_seace_excel(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    require_source_access(current_user, "seace_public_excel")
    suffix = Path(file.filename or "").suffix or ".xls"
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(await file.read())
        tmp_path = Path(tmp.name)
    try:
        df = read_seace_export(tmp_path)
        run = ScrapeRun(
            source="seace_public_excel",
            status="running",
            started_at=datetime.utcnow(),
            diagnostics=f"Archivo importado: {file.filename}",
        )
        db.add(run)
        db.commit()
        db.refresh(run)
        count = upsert_opportunities(db, df, "seace_public_excel", run_id=run.id)
        run.rows_found = count
        run.status = "completed"
        run.finished_at = datetime.utcnow()
        db.commit()
        evaluate_new_opportunity_alerts(db, run.id)
        evaluate_alerts(db)
        if settings.auto_send_alerts:
            send_pending_alerts(db)
        return {"imported": count}
    finally:
        tmp_path.unlink(missing_ok=True)


@router.get("/{opportunity_id}/snapshots", response_model=list[OpportunitySnapshotOut])
def list_snapshots(
    opportunity_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    opportunity = db.get(Opportunity, opportunity_id)
    if not opportunity:
        return []
    require_source_access(current_user, opportunity.source)
    query = (
        select(OpportunitySnapshot)
        .where(OpportunitySnapshot.opportunity_id == opportunity_id)
        .order_by(OpportunitySnapshot.created_at.desc())
        .limit(100)
    )
    return list(db.scalars(query).all())
