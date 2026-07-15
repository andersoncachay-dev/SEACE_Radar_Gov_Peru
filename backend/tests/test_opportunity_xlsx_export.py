from __future__ import annotations

import asyncio
import unittest
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from backend.app.database import Base
from backend.app.routers.opportunities import export_opportunities_xlsx
from backend.app.schemas import OpportunityExcelExportIn


class OpportunityXlsxExportTests(unittest.TestCase):
    def test_generates_modern_xlsx_workbook(self) -> None:
        engine = create_engine("sqlite:///:memory:")
        Base.metadata.create_all(engine)
        db = Session(engine)
        response = export_opportunities_xlsx(
            OpportunityExcelExportIn(
                title="Oportunidades GovRadar",
                headers=["Prioridad", "Proceso"],
                rows=[["A", "CP-ABR-1-2026"]],
            ),
            current_user=object(),
            db=db,
        )

        async def collect() -> bytes:
            return b"".join([chunk async for chunk in response.body_iterator])

        content = asyncio.run(collect())
        workbook = load_workbook(BytesIO(content), read_only=True)

        self.assertTrue(content.startswith(b"PK"))
        self.assertEqual(response.media_type, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertEqual(workbook["Oportunidades"]["B2"].value, "CP-ABR-1-2026")
        self.assertIn("Definición prioridades", workbook.sheetnames)
        self.assertEqual(workbook["Definición prioridades"]["A4"].value, "A")
        self.assertEqual(workbook["Definición prioridades"]["B4"].value, "70 a 100")
        db.close()

    def test_chile_definitions_exclude_peru_only_factors(self) -> None:
        engine = create_engine("sqlite:///:memory:")
        Base.metadata.create_all(engine)
        db = Session(engine)
        response = export_opportunities_xlsx(
            OpportunityExcelExportIn(country="chile", headers=["Prioridad"], rows=[["B"]]),
            current_user=object(), db=db,
        )
        async def collect() -> bytes:
            return b"".join([chunk async for chunk in response.body_iterator])
        workbook = load_workbook(BytesIO(asyncio.run(collect())), read_only=True)
        definitions = workbook["Definición prioridades"]
        values = [[cell.value for cell in row] for row in definitions.iter_rows()]
        entity_row = next(row for row in values if row[0] == "Entidad objetivo")
        quick_row = next(row for row in values if row[0] == "Compra rápida")
        self.assertEqual(entity_row[1], "No aplica")
        self.assertEqual(entity_row[2], "No aplica")
        self.assertEqual(quick_row[1], "No aplica")
        self.assertEqual(quick_row[2], "No aplica")
        db.close()


if __name__ == "__main__":
    unittest.main()
